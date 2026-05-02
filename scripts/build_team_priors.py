"""Bygg team_priors.json fra Scraper-output (Premier League historisk data).

Kilder (sjekkes i denne rekkefølgen):
  1. $TSDL_PL_XLSX miljøvariabel (full sti)
  2. ../../Scraper/output/tsdl_Premier_League.xlsx (relativt mtp utviklingsmaskinen)
  3. Skipp stille hvis ikke funnet

Output: data/team_priors.json med per-lag xG, xGA, clean sheets, 1./2.
omgangs-mønster, BTTS-rater og hjemme/borte-split.

Kjøres:
  - Manuelt: python scripts/build_team_priors.py
  - Auto: scripts/update_priors.ps1 (Windows Task Scheduler, 2× per uke)
  - I CI: kan også fungere uten kilde — committer ikke noe da
"""
from __future__ import annotations

import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl ikke installert. Kjør: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_PATH = REPO_ROOT / "data" / "team_priors.json"


# Ligaer vi støtter. Premier League er primær (FPL = PL), men Bundesliga er
# verdifull tilleggsdata for europeisk kontekst — viser også at appen forstår
# fotball på tvers av store ligaer.
SUPPORTED_LEAGUES = [
    {"key": "PL", "name": "Premier League", "filename": "tsdl_Premier_League.xlsx"},
    {"key": "BL", "name": "Bundesliga", "filename": "tsdl_Bundesliga.xlsx"},
]


def _find_source(filename: str) -> Path | None:
    env = os.environ.get("TSDL_PL_XLSX")
    if env and filename == "tsdl_Premier_League.xlsx":
        p = Path(env)
        if p.exists():
            return p
    candidates = [
        REPO_ROOT.parent / "Scraper" / "output" / filename,
        Path(f"C:/Users/rober/Claude prosjekter/Scraper/output/{filename}"),
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


# Map fra navnene i scraperen til FPL sine offisielle lagnavn så vi kan slå opp
# på FPL-team-ID. Ufullstendig forsetlig — vi normaliserer på lower-strip.
TEAM_NAME_ALIASES = {
    "man city": "Manchester City",
    "man utd": "Manchester United",
    "man united": "Manchester United",
    "newcastle": "Newcastle United",
    "nottm forest": "Nottingham Forest",
    "nott'm forest": "Nottingham Forest",
    "spurs": "Tottenham",
    "tottenham": "Tottenham",
    "wolves": "Wolverhampton",
    "wolverhampton": "Wolverhampton",
    "leicester": "Leicester City",
    "leeds": "Leeds United",
    "west ham": "West Ham",
    "brighton": "Brighton",
    "bournemouth": "Bournemouth",
    "crystal palace": "Crystal Palace",
    "everton": "Everton",
    "fulham": "Fulham",
    "arsenal": "Arsenal",
    "chelsea": "Chelsea",
    "liverpool": "Liverpool",
    "aston villa": "Aston Villa",
    "burnley": "Burnley",
    "sheffield utd": "Sheffield United",
    "sheffield united": "Sheffield United",
    "luton": "Luton",
    "ipswich": "Ipswich",
    "southampton": "Southampton",
    "sunderland": "Sunderland",
    "brentford": "Brentford",
}


def _norm_name(name: str | None) -> str | None:
    if not name:
        return None
    raw = str(name).strip().lower()
    return TEAM_NAME_ALIASES.get(raw, name.strip())


def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace("%", "").replace(",", ".").strip())
    except (TypeError, ValueError):
        return None


def _to_pct(v) -> float | None:
    """Tolker '53%' eller '0.53' eller '53' som en prosent (0.0-1.0)."""
    f = _to_float(v)
    if f is None:
        return None
    return f / 100.0 if f > 1.0 else f


def _extract_xg(wb) -> dict[str, dict]:
    """Henter xG/xGA hjemme/borte fra xg_xG_Against (xG) og xg_xG_Against_2 (xGA)."""
    teams: dict[str, dict] = {}

    if "xg_xG_Against" in wb.sheetnames:
        ws = wb["xg_xG_Against"]
        # Layout: home: G | S | TxG | xG | (sep) | away: G | S | TxG | xG
        # Header row 2: [None, 'G','S','TxG','xG', None, 'G','S','TxG','xG']
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        for r in rows:
            if not r or len(r) < 11:
                continue
            home_team = _norm_name(r[1])
            home_xg = _to_float(r[5])  # xG home
            away_team = _norm_name(r[6])
            away_xg = _to_float(r[10])  # xG away
            if home_team and home_xg is not None:
                teams.setdefault(home_team, {})["xg_home"] = round(home_xg, 2)
            if away_team and away_xg is not None:
                teams.setdefault(away_team, {})["xg_away"] = round(away_xg, 2)

    if "xg_xG_Against_2" in wb.sheetnames:
        ws = wb["xg_xG_Against_2"]
        # xGA layout — column 'C' is conceded count, last col xGA
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        for r in rows:
            if not r or len(r) < 11:
                continue
            home_team = _norm_name(r[1])
            home_xga = _to_float(r[5])
            away_team = _norm_name(r[6])
            away_xga = _to_float(r[10])
            if home_team and home_xga is not None:
                teams.setdefault(home_team, {})["xga_home"] = round(home_xga, 2)
            if away_team and away_xga is not None:
                teams.setdefault(away_team, {})["xga_away"] = round(away_xga, 2)

    return teams


def _extract_clean_sheets(wb) -> dict[str, dict]:
    """Clean-sheet % hjemme/borte fra clean-sheets_*."""
    teams: dict[str, dict] = {}
    sheet = "clean-sheets_Failed_To_Score"  # CS = clean sheets achieved
    if sheet not in wb.sheetnames:
        return teams
    ws = wb[sheet]
    # Layout R2: [None, 'G', 'CS', '%', None, 'G', 'CS', '%', None, 'G', 'CS', '%']
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or len(r) < 12:
            continue
        ht = _norm_name(r[1])
        cs_pct_home = _to_pct(r[4])
        at = _norm_name(r[5])
        cs_pct_away = _to_pct(r[8])
        ot = _norm_name(r[9])
        cs_pct_overall = _to_pct(r[12]) if len(r) > 12 else None
        if ht and cs_pct_home is not None:
            teams.setdefault(ht, {})["cs_pct_home"] = round(cs_pct_home, 3)
        if at and cs_pct_away is not None:
            teams.setdefault(at, {})["cs_pct_away"] = round(cs_pct_away, 3)
        if ot and cs_pct_overall is not None:
            teams.setdefault(ot, {})["cs_pct_overall"] = round(cs_pct_overall, 3)
    return teams


def _extract_half_split(wb) -> dict[str, dict]:
    """1./2.-omgangs scoring-andel."""
    teams: dict[str, dict] = {}
    sheet = "1st-2nd-half-goals_1st_2nd_Half"
    if sheet not in wb.sheetnames:
        return teams
    ws = wb[sheet]
    # R2: [None, 'P', 'TS', '1st H', '1st %', '1st Av', '2nd H', '2nd %', '2nd Av']
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or len(r) < 9:
            continue
        team = _norm_name(r[1])
        first_half_pct = _to_pct(r[5])
        second_half_pct = _to_pct(r[8])
        if team:
            entry = teams.setdefault(team, {})
            if first_half_pct is not None:
                entry["goals_first_half_pct"] = round(first_half_pct, 3)
            if second_half_pct is not None:
                entry["goals_second_half_pct"] = round(second_half_pct, 3)
    return teams


def _extract_btts(wb) -> dict[str, dict]:
    """BTTS-rate per lag."""
    teams: dict[str, dict] = {}
    # Search for the most likely "Quick Table" sheet
    sheet = next((s for s in wb.sheetnames if s.startswith("btts_") and "Quick" in s), None)
    if not sheet:
        return teams
    ws = wb[sheet]
    # Typically layout: [#, rank, team, P, BTTS, %]
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or len(r) < 5:
            continue
        # Find team name and a percent column heuristically
        team = None
        pct = None
        for cell in r[:8]:
            s = str(cell or "").strip()
            if "%" in s and pct is None:
                pct = _to_pct(s)
            elif s and not s.replace(".", "").isdigit() and not team and len(s) > 2:
                team = _norm_name(s)
        if team and pct is not None:
            teams.setdefault(team, {})["btts_pct_overall"] = round(pct, 3)
    return teams


def _extract_form_recent(wb) -> dict[str, dict]:
    """Siste 5 W-D-L sammendrag fra wdl_-ark."""
    teams: dict[str, dict] = {}
    # Use the first wdl Quick Table-ish sheet
    sheet = next((s for s in wb.sheetnames if s.startswith("wdl_")), None)
    if not sheet:
        return teams
    ws = wb[sheet]
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or len(r) < 6:
            continue
        team = None
        for cell in r[:5]:
            s = str(cell or "").strip()
            if s and not s.replace(".", "").isdigit() and len(s) > 2:
                team = _norm_name(s)
                break
        # We don't have a clear schema for wdl form here — skip detailed extraction
        # Could be expanded later with a known schema.
    return teams


def build_priors(source: Path) -> dict:
    print(f"Reading {source} ...")
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)

    teams: dict[str, dict] = {}

    def _merge(source_dict: dict[str, dict]):
        for team, data in source_dict.items():
            if not team:
                continue
            teams.setdefault(team, {}).update(data)

    _merge(_extract_xg(wb))
    _merge(_extract_clean_sheets(wb))
    _merge(_extract_half_split(wb))
    _merge(_extract_btts(wb))

    # Compute derived: combined home+away xG, clean-sheet bias, etc.
    for team, data in teams.items():
        xg_h = data.get("xg_home")
        xg_a = data.get("xg_away")
        if xg_h is not None and xg_a is not None:
            data["xg_avg"] = round((xg_h + xg_a) / 2, 2)
            data["home_advantage_xg"] = round(xg_h - xg_a, 2)
        xga_h = data.get("xga_home")
        xga_a = data.get("xga_away")
        if xga_h is not None and xga_a is not None:
            data["xga_avg"] = round((xga_h + xga_a) / 2, 2)
        # Net xG (attack - defense)
        if data.get("xg_avg") is not None and data.get("xga_avg") is not None:
            data["net_xg"] = round(data["xg_avg"] - data["xga_avg"], 2)

    return {
        "version": 1,
        "source": source.name,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "teams_count": len(teams),
        "teams": teams,
    }


def main() -> int:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    leagues_data = {}
    total = 0

    for spec in SUPPORTED_LEAGUES:
        source = _find_source(spec["filename"])
        if not source:
            print(f"Skipping {spec['name']}: source not found ({spec['filename']}).")
            continue
        priors = build_priors(source)
        priors["league_key"] = spec["key"]
        priors["league_name"] = spec["name"]
        leagues_data[spec["key"]] = priors
        total += priors.get("teams_count", 0)
        sample = next(iter(priors["teams"].items()), None)
        if sample:
            print(f"  {spec['name']} sample {sample[0]}: net_xg={sample[1].get('net_xg')}")

    if not leagues_data:
        print("No leagues processed. Exit.")
        return 0

    output = {
        "version": 2,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "leagues": leagues_data,
        "total_teams": total,
    }
    tmp = OUTPUT_PATH.with_suffix(".json.tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2, sort_keys=True)
    tmp.replace(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH} — {total} teams across {len(leagues_data)} leagues")
    return 0


if __name__ == "__main__":
    sys.exit(main())
